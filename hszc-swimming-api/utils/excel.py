"""Excel 工具类"""
import re
from datetime import datetime, timedelta
from io import BytesIO
import openpyxl


class ExcelHelper:
    """Excel 处理辅助类"""

    def __init__(self, file, sheet_name=None):
        """初始化"""
        self.file = file
        self.workbook = None
        self.sheet = None
        self._load(sheet_name)

    def _load(self, sheet_name=None):
        """加载文件"""
        if hasattr(self.file, 'read'):
            self.workbook = openpyxl.load_workbook(BytesIO(self.file.read()))
        else:
            self.workbook = openpyxl.load_workbook(self.file)

        if sheet_name:
            self.sheet = self.workbook[sheet_name] if sheet_name in self.workbook.sheetnames else self.workbook.active
        else:
            self.sheet = self.workbook.active

    def read_data(self, header_row=1):
        """读取数据为字典列表"""
        headers = []
        data = []

        for row_idx, row in enumerate(self.sheet.iter_rows(values_only=True), 1):
            if row_idx == header_row:
                headers = [str(cell).strip() if cell else '' for cell in row]
            else:
                if any(row):
                    row_data = {}
                    for idx, cell in enumerate(row):
                        if idx < len(headers) and headers[idx]:
                            row_data[headers[idx]] = cell
                    data.append(row_data)

        return data

    def read_one(self, header_row=1):
        """读取第一行数据"""
        data = self.read_data(header_row)
        return data[0] if data else None

    @staticmethod
    def export_to_bytes(headers, data_list):
        """导出数据到 BytesIO"""
        wb = openpyxl.Workbook()
        ws = wb.active

        # 写入表头
        ws.append(headers)

        # 写入数据
        for item in data_list:
            if isinstance(item, dict):
                ws.append([item.get(h, '') for h in headers])
            else:
                ws.append(item)

        # 保存到 BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def export_to_response(headers, data_list, filename='export.xlsx'):
        """导出为 Django HttpResponse """
        from django.http import HttpResponse
        output = ExcelHelper.export_to_bytes(headers, data_list)
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    @staticmethod
    def parse_time_format(value):
        """
        解析时间格式 mm:ss.00 或 mm:ss.f 的成绩
        返回总秒数（decimal）
        """
        if value is None or value == '-':
            return None

        # 如果是数字（openpyxl time对象），直接转换
        if isinstance(value, timedelta):
            return float(value.total_seconds())
        if isinstance(value, datetime):
            return float(value.hour * 3600 + value.minute * 60 + value.second) + value.microsecond / 1000000

        # 如果是字符串，解析格式 mm:ss.00
        if isinstance(value, str):
            value = value.strip()
            # 匹配 mm:ss.ff 或 mm:ss.f 格式
            match = re.match(r'(\d+):(\d{2})\.(\d{1,2})', value)
            if match:
                minutes = int(match.group(1))
                seconds = int(match.group(2))
                fraction = match.group(3)
                # 处理不同位数的小数部分
                if len(fraction) == 1:
                    fraction = int(fraction) * 10
                elif len(fraction) == 2:
                    fraction = int(fraction)
                else:
                    fraction = int(fraction[:2])
                return float(minutes * 60 + seconds) + fraction / 100
            # 尝试匹配 ss.00 格式（秒为单位）
            match = re.match(r'(\d+)\.(\d{1,2})', value)
            if match:
                seconds = int(match.group(1))
                fraction = match.group(2)
                if len(fraction) == 1:
                    fraction = int(fraction) * 10
                else:
                    fraction = int(fraction[:2])
                return float(seconds) + fraction / 100

        # 如果是浮点数，直接返回
        try:
            return float(value)
        except (ValueError, TypeError):
            return None

    @staticmethod
    def format_time_to_string(seconds):
        """将秒数转换为 mm:ss.00 格式字符串"""
        if seconds is None:
            return '-'
        minutes = int(seconds // 60)
        secs = seconds % 60
        return f'{minutes:02d}:{secs:05.2f}'

    @staticmethod
    def get_sheet_names():
        """获取所有sheet名称"""
        return self.workbook.sheetnames if self.workbook else []
